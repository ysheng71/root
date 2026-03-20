"""
Build and format standardized financial statements.

Uses metric_mappings in the DB to resolve XBRL facts to standardized metrics.
Supports annual (FY) and quarterly (Q1-Q4) granularity, multiple output formats.
"""

from __future__ import annotations

import csv
import io
import math
import sqlite3
import sys
from dataclasses import dataclass
from datetime import datetime
from typing import Any, Dict, List, Optional, Tuple

from .metrics import (
    STATEMENT_METRICS,
    STATEMENT_TITLES,
    ALL_METRICS,
    MetricDef,
    get_metric,
)

SCALE_LABELS = {
    1: "USD",
    1_000: "USD in Thousands",
    1_000_000: "USD in Millions",
    1_000_000_000: "USD in Billions",
}

# Section display names for headers
SECTION_HEADERS = {
    # Income Statement
    "revenue": None,          # no explicit header, first section
    "opex": "Operating Expenses",
    "operating": None,
    "below_line": "Non-Operating",
    "bottom_line": None,
    "supplemental": "Supplemental",
    "per_share": "Per Share",
    # Balance Sheet
    "current_assets": "Current Assets",
    "noncurrent_assets": "Non-Current Assets",
    "current_liabilities": "Current Liabilities",
    "noncurrent_liabilities": "Non-Current Liabilities",
    "equity": "Shareholders' Equity",
    # Cash Flow
    "operating": "Operating Activities",
    "investing": "Investing Activities",
    "financing": "Financing Activities",
    "summary": "Summary",
}


@dataclass
class ReportRow:
    metric: MetricDef
    values: Dict[str, Optional[float]]  # period_label -> raw value (unscaled)
    is_section_header: bool = False
    section_label: str = ""
    # Annotation rows: computed metrics injected after their parent metric row
    is_annotation: bool = False
    annotation_display: str = ""   # display name (e.g. "· Revenue Growth (YoY)")
    annotation_fmt: str = ""       # fmt key for _fmt_ratio_value


@dataclass
class Report:
    company_name: str
    ticker: str
    statement: str          # income_statement | balance_sheet | cash_flow
    statement_title: str
    period: str             # annual | quarterly
    periods: List[str]      # ordered period labels, oldest→newest
    rows: List[ReportRow]
    scale: int              # USD divisor (1, 1000, 1_000_000, etc.)


# ---------------------------------------------------------------------------
# Data fetching
# ---------------------------------------------------------------------------

def _fetch_facts(
    conn: sqlite3.Connection,
    cik: str,
    statement: str,
    period: str,
    num_periods: int,
) -> Dict[Tuple[str, str], Optional[float]]:
    """
    Returns {(metric_name, period_label): best_value} for non-derived metrics.

    Uses ROW_NUMBER() to select: lowest concept priority first, then most
    recently filed version (handles restatements).
    """
    if period == "annual":
        form_values = "('10-K', '10-K/A')"
        period_values = "('FY')"
        year_limit = num_periods + 3
        period_key_expr = (
            "printf('%d/%d/%02d',"
            "  cast(strftime('%m', f.period_end) as integer),"
            "  cast(strftime('%d', f.period_end) as integer),"
            "  cast(strftime('%Y', f.period_end) as integer) % 100)"
        )
        sort_col = "period_end"
        sql = f"""
            WITH min_date AS (
                SELECT date(MAX(period_end), '-{year_limit} years') AS cutoff
                FROM xbrl_facts
                WHERE cik = :cik
                  AND form IN {form_values}
            ),
            ranked AS (
                SELECT
                    mm.metric_name,
                    mm.priority    AS concept_priority,
                    {period_key_expr} AS period_label,
                    f.fiscal_year,
                    f.fiscal_period,
                    f.period_end,
                    f.value,
                    f.filed_date,
                    ROW_NUMBER() OVER (
                        PARTITION BY mm.metric_name, f.period_end
                        ORDER BY mm.priority ASC, f.filed_date DESC
                    ) AS rn
                FROM xbrl_facts f
                JOIN metric_mappings mm
                    ON  mm.concept   = f.concept
                    AND mm.taxonomy  = f.taxonomy
                    AND mm.statement = :statement
                    AND f.unit       = mm.unit
                CROSS JOIN min_date
                WHERE f.cik = :cik
                  AND f.form IN {form_values}
                  AND f.fiscal_period IN {period_values}
                  AND f.period_end >= min_date.cutoff
                  AND f.value IS NOT NULL
            )
            SELECT metric_name, period_label, fiscal_year, fiscal_period,
                   period_end, value
            FROM ranked
            WHERE rn = 1
            ORDER BY metric_name, {sort_col}
        """
    else:
        form_values = "('10-Q', '10-Q/A')"
        period_values = "('Q1', 'Q2', 'Q3', 'Q4')"
        year_limit = (num_periods // 4) + 3
        period_key_expr = "printf('%d %s', f.fiscal_year, f.fiscal_period)"
        sort_col = "period_end"
        sql = f"""
            WITH min_year AS (
                SELECT MAX(fiscal_year) - {year_limit} AS cutoff
                FROM xbrl_facts
                WHERE cik = :cik
                  AND form IN {form_values}
            ),
            ranked AS (
                SELECT
                    mm.metric_name,
                    mm.priority    AS concept_priority,
                    {period_key_expr} AS period_label,
                    f.fiscal_year,
                    f.fiscal_period,
                    f.period_end,
                    f.value,
                    f.filed_date,
                    ROW_NUMBER() OVER (
                        PARTITION BY mm.metric_name, f.fiscal_year, f.fiscal_period
                        ORDER BY mm.priority ASC, f.filed_date DESC
                    ) AS rn
                FROM xbrl_facts f
                JOIN metric_mappings mm
                    ON  mm.concept   = f.concept
                    AND mm.taxonomy  = f.taxonomy
                    AND mm.statement = :statement
                    AND f.unit       = mm.unit
                CROSS JOIN min_year
                WHERE f.cik = :cik
                  AND f.form IN {form_values}
                  AND f.fiscal_period IN {period_values}
                  AND f.fiscal_year >= min_year.cutoff
                  AND f.value IS NOT NULL
            )
            SELECT metric_name, period_label, fiscal_year, fiscal_period,
                   period_end, value
            FROM ranked
            WHERE rn = 1
            ORDER BY metric_name, {sort_col}
        """
    cursor = conn.execute(sql, {
        "cik": cik,
        "statement": statement,
    })

    result: Dict[Tuple[str, str], Optional[float]] = {}
    for row in cursor.fetchall():
        key = (row["metric_name"], row["period_label"])
        result[key] = row["value"]

    return result


def _ordered_periods(
    facts: Dict[Tuple[str, str], Optional[float]],
    period: str,
    num_periods: int,
) -> List[str]:
    """Extract sorted period labels from facts, capped at num_periods (most recent)."""
    all_periods = sorted({p for (_, p) in facts.keys()}, key=_period_sort_key)
    return all_periods[-num_periods:]  # keep most recent N


def _period_sort_key(label: str) -> Tuple:
    """Sort key for period labels: 'M/D/YY', 'FY2023', '2023 Q1', or 'LTM'."""
    if label == "LTM":
        return (9999, 0, 0)
    if "/" in label:
        dt = datetime.strptime(label, "%m/%d/%y")
        return (dt.year, dt.month, dt.day)
    if label.startswith("FY"):
        return (int(label[2:]), 0, 0)
    parts = label.split()
    if len(parts) == 2:
        return (int(parts[0]), int(parts[1][1]), 0)
    return (0, 0, 0)


def _compute_derived(
    metric: MetricDef,
    periods: List[str],
    values: Dict[Tuple[str, str], Optional[float]],
) -> Dict[str, Optional[float]]:
    """Evaluate derived_expr for each period."""
    result: Dict[str, Optional[float]] = {}
    for p in periods:
        try:
            # Build a namespace of {metric_name: value} for this period
            ns: Dict[str, float] = {}
            for m in ALL_METRICS:
                v = values.get((m.name, p))
                ns[m.name] = v if v is not None else float("nan")
            val = eval(metric.derived_expr, {"__builtins__": {}}, ns)  # noqa: S307
            import math
            result[p] = None if (math.isnan(val) or math.isinf(val)) else val
        except Exception:
            result[p] = None
    return result


def _compute_derived_single(
    metric: MetricDef,
    period: str,
    values: Dict[Tuple[str, str], Optional[float]],
) -> Optional[float]:
    """Evaluate derived_expr for a single period using flat (name, period) facts dict."""
    try:
        ns: Dict[str, float] = {}
        for m in ALL_METRICS:
            v = values.get((m.name, period))
            ns[m.name] = v if v is not None else float("nan")
        val = eval(metric.derived_expr, {"__builtins__": {}}, ns)  # noqa: S307
        return None if (math.isnan(val) or math.isinf(val)) else val
    except Exception:
        return None


def _eval_derived_for_period(
    metric: MetricDef,
    period: str,
    pool: Dict[str, Dict[str, Optional[float]]],
) -> Optional[float]:
    """Evaluate derived_expr for a single period using pool {name: {period: value}}."""
    try:
        ns = {
            m.name: (
                pool.get(m.name, {}).get(period)
                if pool.get(m.name, {}).get(period) is not None
                else float("nan")
            )
            for m in ALL_METRICS
        }
        val = eval(metric.derived_expr, {"__builtins__": {}}, ns)  # noqa: S307
        return None if (math.isnan(val) or math.isinf(val)) else val
    except Exception:
        return None


def _fetch_ltm(
    conn: sqlite3.Connection,
    cik: str,
    statement: str,
) -> Dict[str, Optional[float]]:
    """
    Compute LTM (Last Twelve Months) values for each metric in a statement.

    For duration metrics: annual + recent_ytd - prior_ytd
    For instant metrics: most recent quarterly snapshot (fallback to annual)
    """
    # Q1: most recent annual value per metric
    q1_sql = """
        WITH ranked AS (
            SELECT mm.metric_name, mm.period_type, f.period_end, f.value,
                   ROW_NUMBER() OVER (
                       PARTITION BY mm.metric_name
                       ORDER BY f.period_end DESC, f.filed_date DESC
                   ) AS rn
            FROM xbrl_facts f
            JOIN metric_mappings mm
                ON mm.concept = f.concept AND mm.taxonomy = f.taxonomy
                   AND mm.statement = :stmt AND f.unit = mm.unit
            WHERE f.cik = :cik AND f.form IN ('10-K', '10-K/A')
              AND f.fiscal_period = 'FY'
              AND f.value IS NOT NULL
        )
        SELECT metric_name, period_type, period_end, value FROM ranked WHERE rn = 1
    """
    rows = conn.execute(q1_sql, {"cik": cik, "stmt": statement}).fetchall()
    if not rows:
        return {}

    annual: Dict[str, float] = {}
    annual_period_type: Dict[str, str] = {}
    annual_end: Optional[str] = None
    for row in rows:
        annual[row["metric_name"]] = row["value"]
        annual_period_type[row["metric_name"]] = row["period_type"]
        if annual_end is None or row["period_end"] > annual_end:
            annual_end = row["period_end"]

    # Q2: most recent quarterly YTD for duration metrics (period_end > annual_end)
    q2_sql = """
        WITH ranked AS (
            SELECT mm.metric_name, f.period_end, f.value,
                   ROW_NUMBER() OVER (
                       PARTITION BY mm.metric_name
                       ORDER BY f.period_end DESC, f.filed_date DESC
                   ) AS rn
            FROM xbrl_facts f
            JOIN metric_mappings mm
                ON mm.concept = f.concept AND mm.taxonomy = f.taxonomy
                   AND mm.statement = :stmt AND f.unit = mm.unit
            WHERE f.cik = :cik AND f.form IN ('10-Q', '10-Q/A')
              AND f.fiscal_period IN ('Q1', 'Q2', 'Q3', 'Q4')
              AND mm.period_type = 'duration'
              AND f.period_end > :annual_end
              AND f.value IS NOT NULL
        )
        SELECT metric_name, period_end, value FROM ranked WHERE rn = 1
    """
    q2_rows = conn.execute(q2_sql, {
        "cik": cik, "stmt": statement, "annual_end": annual_end,
    }).fetchall()

    recent_ytd: Dict[str, float] = {}
    recent_q_end: Optional[str] = None
    prior_ytd: Dict[str, float] = {}

    if q2_rows:
        for row in q2_rows:
            recent_ytd[row["metric_name"]] = row["value"]
            if recent_q_end is None or row["period_end"] > recent_q_end:
                recent_q_end = row["period_end"]

        # Compute prior_q_end = one year before recent_q_end
        rq = datetime.strptime(recent_q_end, "%Y-%m-%d")
        try:
            prior_q = rq.replace(year=rq.year - 1)
        except ValueError:
            prior_q = rq.replace(year=rq.year - 1, day=28)
        prior_q_end = prior_q.strftime("%Y-%m-%d")

        # Q3: prior-year same quarter YTD for duration metrics
        q3_sql = """
            WITH ranked AS (
                SELECT mm.metric_name, f.value,
                       ROW_NUMBER() OVER (
                           PARTITION BY mm.metric_name
                           ORDER BY f.filed_date DESC
                       ) AS rn
                FROM xbrl_facts f
                JOIN metric_mappings mm
                    ON mm.concept = f.concept AND mm.taxonomy = f.taxonomy
                       AND mm.statement = :stmt AND f.unit = mm.unit
                WHERE f.cik = :cik AND f.form IN ('10-Q', '10-Q/A')
                  AND f.fiscal_period IN ('Q1', 'Q2', 'Q3', 'Q4')
                  AND mm.period_type = 'duration'
                  AND f.period_end = :prior_q_end
                  AND f.value IS NOT NULL
            )
            SELECT metric_name, value FROM ranked WHERE rn = 1
        """
        q3_rows = conn.execute(q3_sql, {
            "cik": cik, "stmt": statement, "prior_q_end": prior_q_end,
        }).fetchall()
        prior_ytd = {row["metric_name"]: row["value"] for row in q3_rows}

    # Q4: most recent quarterly snapshot for instant metrics (period_end > annual_end)
    q4_sql = """
        WITH ranked AS (
            SELECT mm.metric_name, f.value,
                   ROW_NUMBER() OVER (
                       PARTITION BY mm.metric_name
                       ORDER BY f.period_end DESC, f.filed_date DESC
                   ) AS rn
            FROM xbrl_facts f
            JOIN metric_mappings mm
                ON mm.concept = f.concept AND mm.taxonomy = f.taxonomy
                   AND mm.statement = :stmt AND f.unit = mm.unit
            WHERE f.cik = :cik AND f.form IN ('10-Q', '10-Q/A')
              AND f.fiscal_period IN ('Q1', 'Q2', 'Q3', 'Q4')
              AND mm.period_type = 'instant'
              AND f.period_end > :annual_end
              AND f.value IS NOT NULL
        )
        SELECT metric_name, value FROM ranked WHERE rn = 1
    """
    q4_rows = conn.execute(q4_sql, {
        "cik": cik, "stmt": statement, "annual_end": annual_end,
    }).fetchall()
    snapshot: Dict[str, float] = {row["metric_name"]: row["value"] for row in q4_rows}

    # Assemble LTM values
    ltm: Dict[str, Optional[float]] = {}
    for name, ann_val in annual.items():
        ptype = annual_period_type.get(name, "duration")
        if ptype == "instant":
            ltm[name] = snapshot.get(name, ann_val)
        else:
            recent = recent_ytd.get(name)
            prior = prior_ytd.get(name)
            if recent is not None and prior is not None:
                ltm[name] = ann_val + recent - prior
            else:
                ltm[name] = ann_val

    return ltm


# ---------------------------------------------------------------------------
# Report builder
# ---------------------------------------------------------------------------

def build_report(
    conn: sqlite3.Connection,
    cik: str,
    ticker: str,
    company_name: str,
    statement: str,
    period: str,
    num_periods: int,
    scale: int = 1_000_000,
    annotation_pool: Optional[Dict[str, Dict[str, Optional[float]]]] = None,
) -> Report:
    from .annotation_defs import get_annotations
    from .ratio_defs import ALL_COMPUTED

    metrics = STATEMENT_METRICS[statement]

    # Fetch raw facts
    facts = _fetch_facts(conn, cik, statement, period, num_periods)

    # Determine periods present in data
    periods = _ordered_periods(facts, period, num_periods)

    # Compute derived metrics and merge into facts dict
    for m in metrics:
        if m.is_derived and m.derived_expr:
            derived_vals = _compute_derived(m, periods, facts)
            for p, v in derived_vals.items():
                facts[(m.name, p)] = v

    # Add LTM column for annual reports
    if period == "annual":
        ltm_vals = _fetch_ltm(conn, cik, statement)
        for name, val in ltm_vals.items():
            facts[(name, "LTM")] = val
        for m in metrics:
            if m.is_derived and m.derived_expr:
                facts[(m.name, "LTM")] = _compute_derived_single(m, "LTM", facts)
        periods = periods + ["LTM"]

    # Build annotation metadata lookup: name → (display, fmt)
    ann_meta: Dict[str, tuple] = {m.name: (m.display, m.fmt) for m in ALL_COMPUTED}
    annotations = get_annotations(statement, period) if annotation_pool else {}

    # Build rows with section headers
    rows: List[ReportRow] = []
    prev_section = None
    for m in metrics:
        if m.section != prev_section:
            hdr = SECTION_HEADERS.get(m.section)
            if hdr:
                rows.append(ReportRow(
                    metric=m,
                    values={},
                    is_section_header=True,
                    section_label=hdr,
                ))
            prev_section = m.section

        row_values = {p: facts.get((m.name, p)) for p in periods}
        rows.append(ReportRow(metric=m, values=row_values))

        # Inject annotation rows immediately after this metric
        for ann_name in annotations.get(m.name, []):
            display, fmt = ann_meta.get(ann_name, (ann_name, "raw"))
            ann_vals = {p: annotation_pool.get(ann_name, {}).get(p) for p in periods}
            rows.append(ReportRow(
                metric=m,
                values=ann_vals,
                is_annotation=True,
                annotation_display=display,
                annotation_fmt=fmt,
            ))

    return Report(
        company_name=company_name,
        ticker=ticker,
        statement=statement,
        statement_title=STATEMENT_TITLES[statement],
        period=period,
        periods=periods,
        rows=rows,
        scale=scale,
    )


def _get_company(conn: sqlite3.Connection, ticker: str) -> Optional[dict]:
    cursor = conn.execute(
        "SELECT cik, ticker, name FROM companies WHERE ticker = ?",
        (ticker.upper(),),
    )
    row = cursor.fetchone()
    return dict(row) if row else None


def build_reports(
    conn: sqlite3.Connection,
    ticker: str,
    statements: List[str],
    period: str,
    num_periods: int,
    scale: int = 1_000_000,
    price: Optional[float] = None,
) -> List[Report]:
    from .computed import RatioEngine
    from .ratio_defs import ALL_COMPUTED

    company = _get_company(conn, ticker)
    if not company:
        raise ValueError(f"Ticker '{ticker}' not found in database. Run 'fetch' first.")

    # Build unified metric pool once — shared by annotations and the ratio report
    full_pool, ann_periods = fetch_all_metrics(conn, company["cik"], period, num_periods)
    engine = RatioEngine()
    annotation_pool = engine.compute_all(ALL_COMPUTED, full_pool, ann_periods, price=price)

    reports = []
    for stmt in statements:
        if stmt == "ratios":
            r = build_ratio_report(
                conn=conn,
                cik=company["cik"],
                ticker=company["ticker"],
                company_name=company["name"],
                period=period,
                num_periods=num_periods,
                scale=scale,
                price=price,
                prebuilt_pool=annotation_pool,
                prebuilt_periods=ann_periods,
            )
        else:
            r = build_report(
                conn=conn,
                cik=company["cik"],
                ticker=company["ticker"],
                company_name=company["name"],
                statement=stmt,
                period=period,
                num_periods=num_periods,
                scale=scale,
                annotation_pool=annotation_pool,
            )
        reports.append(r)
    return reports


# ---------------------------------------------------------------------------
# Formatting helpers
# ---------------------------------------------------------------------------

def _fmt_value(value: Optional[float], unit: str, scale: int) -> str:
    if value is None:
        return "—"
    if unit == "USD":
        scaled = value / scale
        if abs(scaled) >= 1000:
            return f"{scaled:,.0f}"
        return f"{scaled:,.1f}"
    if unit == "shares":
        scaled = value / 1_000_000
        return f"{scaled:,.1f}M"
    if unit == "USD/shares":
        return f"{value:.2f}"
    return f"{value:,.2f}"


def _scale_label(scale: int) -> str:
    return SCALE_LABELS.get(scale, f"÷{scale:,}")


# ---------------------------------------------------------------------------
# Text formatter
# ---------------------------------------------------------------------------

def format_text(report: Report) -> str:
    LABEL_WIDTH = 42
    COL_WIDTH = 12
    total_width = LABEL_WIDTH + COL_WIDTH * len(report.periods) + 2

    lines: List[str] = []
    sep = "─" * total_width
    thick = "═" * total_width

    # Title block
    lines.append(thick)
    lines.append(
        f"  {report.company_name} ({report.ticker})  —  {report.statement_title}"
    )
    granularity = "Annual" if report.period == "annual" else "Quarterly"
    lines.append(f"  {granularity}  |  {_scale_label(report.scale)}")
    lines.append(thick)

    # Column headers
    header = " " * LABEL_WIDTH
    for p in report.periods:
        header += p.rjust(COL_WIDTH)
    lines.append(header)
    lines.append(sep)

    for row in report.rows:
        if row.is_section_header:
            lines.append("")
            lines.append(f"  {row.section_label.upper()}")
            lines.append(sep)
            continue

        if row.is_annotation:
            label = ("    · " + row.annotation_display)[:LABEL_WIDTH].ljust(LABEL_WIDTH)
            vals = ""
            for p in report.periods:
                vals += _fmt_ratio_value(
                    row.values.get(p), row.annotation_fmt, report.scale
                ).rjust(COL_WIDTH)
            lines.append(label + vals)
            continue

        m = row.metric
        prefix = "  " + ("  " if m.indent else "")
        label = prefix + m.display
        label = label[:LABEL_WIDTH]
        label = label.ljust(LABEL_WIDTH)

        vals = ""
        for p in report.periods:
            vals += _fmt_value(row.values.get(p), m.unit, report.scale).rjust(COL_WIDTH)

        lines.append(label + vals)

        # Add blank line after major totals for readability
        if m.indent == 0 and m.section not in ("revenue",):
            lines.append("")

    lines.append(thick)
    return "\n".join(lines)


# ---------------------------------------------------------------------------
# CSV formatter
# ---------------------------------------------------------------------------

def format_csv(report: Report) -> str:
    buf = io.StringIO()
    writer = csv.writer(buf)

    # Header rows
    writer.writerow(["Company", report.company_name])
    writer.writerow(["Ticker", report.ticker])
    writer.writerow(["Statement", report.statement_title])
    writer.writerow(["Period", "Annual" if report.period == "annual" else "Quarterly"])
    writer.writerow(["Scale", _scale_label(report.scale)])
    writer.writerow([])

    col_headers = ["Metric", "Metric ID", "Unit"] + list(report.periods)
    writer.writerow(col_headers)

    for row in report.rows:
        if row.is_section_header:
            writer.writerow([f"--- {row.section_label} ---"])
            continue
        if row.is_annotation:
            raw_vals = [row.values.get(p) for p in report.periods]
            writer.writerow(
                ["· " + row.annotation_display, "", row.annotation_fmt]
                + ["" if v is None else round(v, 4) for v in raw_vals]
            )
            continue
        m = row.metric
        raw_vals = [row.values.get(p) for p in report.periods]
        # Scale USD values, keep others raw
        scaled_vals = []
        for v in raw_vals:
            if v is None:
                scaled_vals.append("")
            elif m.unit == "USD":
                scaled_vals.append(round(v / report.scale, 2))
            else:
                scaled_vals.append(v)
        writer.writerow([m.display, m.name, m.unit] + scaled_vals)

    return buf.getvalue()


# ---------------------------------------------------------------------------
# Excel formatter
# ---------------------------------------------------------------------------

def write_excel(reports: List[Report], output_path: str) -> None:
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise ImportError(
            "openpyxl is required for Excel export. "
            "Install with: pip install openpyxl"
        )

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default sheet

    HEADER_FILL = PatternFill("solid", fgColor="1F3864")
    HEADER_FONT = Font(color="FFFFFF", bold=True, size=10)
    SECTION_FILL = PatternFill("solid", fgColor="D9E1F2")
    SECTION_FONT = Font(bold=True, size=9, color="1F3864")
    TOTAL_FONT = Font(bold=True, size=9)
    DETAIL_FONT = Font(size=9)
    TITLE_FONT = Font(bold=True, size=12)
    SUBTITLE_FONT = Font(size=9, italic=True)
    ALT_FILL = PatternFill("solid", fgColor="F5F7FA")
    THIN = Side(style="thin", color="D0D0D0")
    THIN_BORDER = Border(bottom=Side(style="thin", color="CCCCCC"))

    for report in reports:
        sheet_name = report.statement_title[:31]
        ws = wb.create_sheet(title=sheet_name)

        # Sheet title
        ws["A1"] = f"{report.company_name} ({report.ticker})"
        ws["A1"].font = TITLE_FONT
        granularity = "Annual" if report.period == "annual" else "Quarterly"
        ws["A2"] = f"{report.statement_title}  ·  {granularity}  ·  {_scale_label(report.scale)}"
        ws["A2"].font = SUBTITLE_FONT
        ws.append([])  # blank row 3

        # Column headers: row 4
        header_row = ["", "Metric ID", "Unit"] + list(report.periods)
        ws.append(header_row)
        hdr_row_num = ws.max_row
        for col_idx, val in enumerate(header_row, 1):
            cell = ws.cell(row=hdr_row_num, column=col_idx)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(
                horizontal="right" if col_idx > 3 else "left",
                vertical="center",
            )
        ws.freeze_panes = ws.cell(row=hdr_row_num + 1, column=4)

        # Data rows
        data_row_num = hdr_row_num
        alt = False
        for row in report.rows:
            data_row_num += 1
            if row.is_section_header:
                ws.cell(row=data_row_num, column=1, value=row.section_label)
                for c in range(1, 4 + len(report.periods)):
                    cell = ws.cell(row=data_row_num, column=c)
                    cell.fill = SECTION_FILL
                    cell.font = SECTION_FONT
                alt = False
                continue

            m = row.metric
            indent_str = "    " * m.indent
            label_cell = ws.cell(row=data_row_num, column=1, value=indent_str + m.display)
            id_cell = ws.cell(row=data_row_num, column=2, value=m.name)
            unit_cell = ws.cell(row=data_row_num, column=3, value=m.unit)

            row_font = TOTAL_FONT if m.indent == 0 else DETAIL_FONT
            label_cell.font = row_font
            id_cell.font = DETAIL_FONT
            unit_cell.font = DETAIL_FONT

            if alt:
                for c in range(1, 4 + len(report.periods)):
                    ws.cell(row=data_row_num, column=c).fill = ALT_FILL

            for col_offset, p in enumerate(report.periods):
                v = row.values.get(p)
                col_num = 4 + col_offset
                cell = ws.cell(row=data_row_num, column=col_num)
                if v is not None:
                    if m.unit == "USD":
                        cell.value = round(v / report.scale, 2)
                        cell.number_format = '#,##0.0'
                    elif m.unit == "shares":
                        cell.value = round(v / 1_000_000, 2)
                        cell.number_format = '#,##0.0'
                    elif m.unit == "USD/shares":
                        cell.value = round(v, 2)
                        cell.number_format = '0.00'
                    else:
                        cell.value = v
                        cell.number_format = '#,##0.00'
                cell.font = row_font
                cell.alignment = Alignment(horizontal="right")

            alt = not alt

        # Column widths
        ws.column_dimensions["A"].width = 38
        ws.column_dimensions["B"].width = 30
        ws.column_dimensions["C"].width = 12
        for col_offset in range(len(report.periods)):
            col_letter = get_column_letter(4 + col_offset)
            ws.column_dimensions[col_letter].width = 14

    wb.save(output_path)


# ---------------------------------------------------------------------------
# Ratio report — data layer
# ---------------------------------------------------------------------------

def _pool_from_flat(
    flat: Dict[Tuple[str, str], Optional[float]],
    periods: List[str],
) -> Dict[str, Dict[str, Optional[float]]]:
    """Reshape flat {(name, period): value} → {name: {period: value}}."""
    pool: Dict[str, Dict[str, Optional[float]]] = {}
    for (name, p), v in flat.items():
        if name not in pool:
            pool[name] = {pp: None for pp in periods}
        pool[name][p] = v
    # Fill missing periods with None
    for name in pool:
        for p in periods:
            pool[name].setdefault(p, None)
    return pool


def fetch_all_metrics(
    conn: sqlite3.Connection,
    cik: str,
    period: str,
    num_periods: int,
) -> Tuple[Dict[str, Dict[str, Optional[float]]], List[str]]:
    """
    Fetch base metrics from all three statements and compute derived metrics.
    Returns (data_pool, sorted_periods).
    data_pool: {metric_name: {period_label: value}}
    """
    # Collect flat facts across all statements
    all_flat: Dict[Tuple[str, str], Optional[float]] = {}
    for stmt in ("income_statement", "balance_sheet", "cash_flow"):
        flat = _fetch_facts(conn, cik, stmt, period, num_periods)
        all_flat.update(flat)

    # Determine the unified period list
    all_periods = _ordered_periods(all_flat, period, num_periods)

    pool = _pool_from_flat(all_flat, all_periods)

    # Compute derived metrics from metrics.py (ebitda, free_cash_flow, etc.)
    # _compute_derived expects {(name, period): value}; adapt from pool
    for m in ALL_METRICS:
        if m.is_derived and m.derived_expr:
            result: Dict[str, Optional[float]] = {}
            for p in all_periods:
                try:
                    ns = {
                        metric.name: (
                            pool.get(metric.name, {}).get(p)
                            if pool.get(metric.name, {}).get(p) is not None
                            else float("nan")
                        )
                        for metric in ALL_METRICS
                    }
                    val = eval(m.derived_expr, {"__builtins__": {}}, ns)  # noqa: S307
                    result[p] = None if (math.isnan(val) or math.isinf(val)) else val
                except Exception:
                    result[p] = None
            pool[m.name] = result

    # Add LTM column for annual reports
    if period == "annual":
        for stmt in ("income_statement", "balance_sheet", "cash_flow"):
            for name, val in _fetch_ltm(conn, cik, stmt).items():
                all_flat[(name, "LTM")] = val
        all_periods = all_periods + ["LTM"]
        pool = _pool_from_flat(all_flat, all_periods)
        for m in ALL_METRICS:
            if m.is_derived and m.derived_expr:
                pool.setdefault(m.name, {})["LTM"] = _eval_derived_for_period(
                    m, "LTM", pool
                )

    return pool, all_periods


# ---------------------------------------------------------------------------
# Ratio report — model
# ---------------------------------------------------------------------------

@dataclass
class RatioRow:
    metric_name: str
    display: str
    fmt: str           # "percent" | "times" | "multiple" | "days" | "currency" | "currency_per_share" | "raw"
    indent: int
    values: Dict[str, Optional[float]]
    is_section_header: bool = False
    section_label: str = ""


@dataclass
class RatioReport:
    company_name: str
    ticker: str
    statement: str = "ratios"
    statement_title: str = "Ratios & Valuation"
    period: str = "annual"
    periods: List[str] = None
    rows: List[RatioRow] = None
    scale: int = 1_000_000
    price: Optional[float] = None


def build_ratio_report(
    conn: sqlite3.Connection,
    cik: str,
    ticker: str,
    company_name: str,
    period: str,
    num_periods: int,
    scale: int = 1_000_000,
    price: Optional[float] = None,
    prebuilt_pool: Optional[Dict[str, Dict[str, Optional[float]]]] = None,
    prebuilt_periods: Optional[List[str]] = None,
) -> RatioReport:
    from .computed import RatioEngine
    from .ratio_defs import ALL_COMPUTED, VISIBLE_COMPUTED, SECTION_TITLES, SECTION_ORDER

    if prebuilt_pool is not None and prebuilt_periods is not None:
        pool = prebuilt_pool
        periods = prebuilt_periods
    else:
        # Build base data pool
        pool, periods = fetch_all_metrics(conn, cik, period, num_periods)
        # Run the ratio engine (evaluates ALL_COMPUTED in order, growing the pool)
        engine = RatioEngine()
        pool = engine.compute_all(ALL_COMPUTED, pool, periods, price=price)

    # Build rows grouped by section
    rows: List[RatioRow] = []
    prev_section = None

    for m in VISIBLE_COMPUTED:
        # Section header
        if m.section != prev_section:
            title = SECTION_TITLES.get(m.section, m.section.replace("_", " ").title())
            # Skip market section header if no price (all dashes anyway — still show it)
            rows.append(RatioRow(
                metric_name="",
                display="",
                fmt="",
                indent=0,
                values={},
                is_section_header=True,
                section_label=title,
            ))
            prev_section = m.section

        vals = {p: pool.get(m.name, {}).get(p) for p in periods}
        rows.append(RatioRow(
            metric_name=m.name,
            display=m.display,
            fmt=m.fmt,
            indent=m.indent,
            values=vals,
        ))

    return RatioReport(
        company_name=company_name,
        ticker=ticker,
        periods=periods,
        rows=rows,
        scale=scale,
        price=price,
        period=period,
    )


# ---------------------------------------------------------------------------
# Ratio value formatter
# ---------------------------------------------------------------------------

def _fmt_ratio_value(value: Optional[float], fmt: str, scale: int) -> str:
    if value is None:
        return "—"
    if fmt == "percent":
        # Values are in percentage points (e.g. 23.4 means 23.4%)
        if abs(value) < 1000:
            return f"{value:.1f}%"
        return f"{value:.0f}%"
    if fmt in ("times", "multiple"):
        return f"{value:.1f}x"
    if fmt == "days":
        return f"{value:.0f}"
    if fmt == "currency":
        scaled = value / scale
        if abs(scaled) >= 1000:
            return f"{scaled:,.0f}"
        return f"{scaled:,.1f}"
    if fmt == "currency_per_share":
        return f"${value:.2f}"
    return f"{value:.2f}"


# ---------------------------------------------------------------------------
# Ratio text formatter
# ---------------------------------------------------------------------------

def format_ratio_text(report: RatioReport) -> str:
    LABEL_WIDTH = 42
    COL_WIDTH = 12
    total_width = LABEL_WIDTH + COL_WIDTH * len(report.periods) + 2

    lines: List[str] = []
    sep = "─" * total_width
    thick = "═" * total_width

    lines.append(thick)
    price_str = f"  |  Price ${report.price:,.2f}" if report.price else ""
    lines.append(f"  {report.company_name} ({report.ticker})  —  {report.statement_title}")
    granularity = "Annual" if report.period == "annual" else "Quarterly"
    lines.append(f"  {granularity}  |  {_scale_label(report.scale)}{price_str}")
    lines.append(thick)

    header = " " * LABEL_WIDTH
    for p in report.periods:
        header += p.rjust(COL_WIDTH)
    lines.append(header)
    lines.append(sep)

    for row in report.rows:
        if row.is_section_header:
            lines.append("")
            lines.append(f"  {row.section_label.upper()}")
            lines.append(sep)
            continue

        prefix = "  " + ("  " if row.indent else "")
        label = (prefix + row.display)[:LABEL_WIDTH].ljust(LABEL_WIDTH)

        vals = ""
        for p in report.periods:
            vals += _fmt_ratio_value(row.values.get(p), row.fmt, report.scale).rjust(COL_WIDTH)

        lines.append(label + vals)
        if row.indent == 0:
            lines.append("")

    lines.append(thick)
    return "\n".join(lines)


# ---------------------------------------------------------------------------
# Ratio CSV formatter
# ---------------------------------------------------------------------------

def format_ratio_csv(report: RatioReport) -> str:
    buf = io.StringIO()
    writer = csv.writer(buf)

    writer.writerow(["Company", report.company_name])
    writer.writerow(["Ticker", report.ticker])
    writer.writerow(["Statement", report.statement_title])
    writer.writerow(["Period", "Annual" if report.period == "annual" else "Quarterly"])
    if report.price:
        writer.writerow(["Price", report.price])
    writer.writerow([])

    writer.writerow(["Metric", "Metric ID", "Format"] + list(report.periods))

    for row in report.rows:
        if row.is_section_header:
            writer.writerow([f"--- {row.section_label} ---"])
            continue
        raw_vals = [row.values.get(p) for p in report.periods]
        writer.writerow([row.display, row.metric_name, row.fmt] + [
            "" if v is None else round(v, 4) for v in raw_vals
        ])

    return buf.getvalue()


# ---------------------------------------------------------------------------
# Patch write_excel to also handle RatioReport
# ---------------------------------------------------------------------------

def _write_ratio_sheet(ws, report: RatioReport, styles: dict) -> None:
    """Write a RatioReport into an existing openpyxl worksheet."""
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    HEADER_FILL = styles["HEADER_FILL"]
    HEADER_FONT = styles["HEADER_FONT"]
    SECTION_FILL = styles["SECTION_FILL"]
    SECTION_FONT = styles["SECTION_FONT"]
    TOTAL_FONT = styles["TOTAL_FONT"]
    DETAIL_FONT = styles["DETAIL_FONT"]
    ALT_FILL = styles["ALT_FILL"]
    TITLE_FONT = styles["TITLE_FONT"]
    SUBTITLE_FONT = styles["SUBTITLE_FONT"]

    price_str = f"  ·  Price ${report.price:,.2f}" if report.price else ""
    ws["A1"] = f"{report.company_name} ({report.ticker})"
    ws["A1"].font = TITLE_FONT
    granularity = "Annual" if report.period == "annual" else "Quarterly"
    ws["A2"] = f"{report.statement_title}  ·  {granularity}{price_str}"
    ws["A2"].font = SUBTITLE_FONT
    ws.append([])

    header_row = ["", "Metric ID", "Format"] + list(report.periods)
    ws.append(header_row)
    hdr_row_num = ws.max_row
    for col_idx, _ in enumerate(header_row, 1):
        cell = ws.cell(row=hdr_row_num, column=col_idx)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="right" if col_idx > 3 else "left")
    ws.freeze_panes = ws.cell(row=hdr_row_num + 1, column=4)

    data_row_num = hdr_row_num
    alt = False
    for row in report.rows:
        data_row_num += 1
        if row.is_section_header:
            ws.cell(row=data_row_num, column=1, value=row.section_label)
            for c in range(1, 4 + len(report.periods)):
                cell = ws.cell(row=data_row_num, column=c)
                cell.fill = SECTION_FILL
                cell.font = SECTION_FONT
            alt = False
            continue

        indent_str = "    " * row.indent
        ws.cell(row=data_row_num, column=1, value=indent_str + row.display)
        ws.cell(row=data_row_num, column=2, value=row.metric_name)
        ws.cell(row=data_row_num, column=3, value=row.fmt)
        row_font = TOTAL_FONT if row.indent == 0 else DETAIL_FONT
        ws.cell(row=data_row_num, column=1).font = row_font

        if alt:
            for c in range(1, 4 + len(report.periods)):
                ws.cell(row=data_row_num, column=c).fill = ALT_FILL

        for col_offset, p in enumerate(report.periods):
            v = row.values.get(p)
            col_num = 4 + col_offset
            cell = ws.cell(row=data_row_num, column=col_num)
            if v is not None:
                cell.value = round(v, 4)
                if row.fmt == "percent":
                    cell.value = round(v / 100, 4)  # Excel % format expects 0-1 fraction
                    cell.number_format = '0.0%'
                elif row.fmt in ("times", "multiple"):
                    cell.number_format = '0.0x'
                elif row.fmt == "days":
                    cell.number_format = '0'
                elif row.fmt == "currency":
                    cell.value = round(v / report.scale, 2)
                    cell.number_format = '#,##0.0'
                elif row.fmt == "currency_per_share":
                    cell.number_format = '$#,##0.00'
            cell.font = row_font
            cell.alignment = Alignment(horizontal="right")

        alt = not alt

    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 12
    for col_offset in range(len(report.periods)):
        ws.column_dimensions[get_column_letter(4 + col_offset)].width = 14


def write_excel(reports, output_path: str) -> None:
    """Write a list of Report and/or RatioReport objects to an Excel workbook."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise ImportError(
            "openpyxl is required for Excel export. Install with: pip install openpyxl"
        )

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    styles = {
        "HEADER_FILL": PatternFill("solid", fgColor="1F3864"),
        "HEADER_FONT": Font(color="FFFFFF", bold=True, size=10),
        "SECTION_FILL": PatternFill("solid", fgColor="D9E1F2"),
        "SECTION_FONT": Font(bold=True, size=9, color="1F3864"),
        "TOTAL_FONT": Font(bold=True, size=9),
        "DETAIL_FONT": Font(size=9),
        "TITLE_FONT": Font(bold=True, size=12),
        "SUBTITLE_FONT": Font(size=9, italic=True),
        "ALT_FILL": PatternFill("solid", fgColor="F5F7FA"),
    }

    for report in reports:
        sheet_name = report.statement_title[:31]
        ws = wb.create_sheet(title=sheet_name)

        if isinstance(report, RatioReport):
            _write_ratio_sheet(ws, report, styles)
            continue

        # --- Original Report sheets (unchanged logic, refactored to use styles dict) ---
        ws["A1"] = f"{report.company_name} ({report.ticker})"
        ws["A1"].font = styles["TITLE_FONT"]
        granularity = "Annual" if report.period == "annual" else "Quarterly"
        ws["A2"] = f"{report.statement_title}  ·  {granularity}  ·  {_scale_label(report.scale)}"
        ws["A2"].font = styles["SUBTITLE_FONT"]
        ws.append([])

        header_row = ["", "Metric ID", "Unit"] + list(report.periods)
        ws.append(header_row)
        hdr_row_num = ws.max_row
        for col_idx, _ in enumerate(header_row, 1):
            cell = ws.cell(row=hdr_row_num, column=col_idx)
            cell.fill = styles["HEADER_FILL"]
            cell.font = styles["HEADER_FONT"]
            cell.alignment = Alignment(horizontal="right" if col_idx > 3 else "left")
        ws.freeze_panes = ws.cell(row=hdr_row_num + 1, column=4)

        ANNOT_FONT = Font(size=9, italic=True, color="555555")

        data_row_num = hdr_row_num
        alt = False
        for row in report.rows:
            data_row_num += 1
            if row.is_section_header:
                ws.cell(row=data_row_num, column=1, value=row.section_label)
                for c in range(1, 4 + len(report.periods)):
                    cell = ws.cell(row=data_row_num, column=c)
                    cell.fill = styles["SECTION_FILL"]
                    cell.font = styles["SECTION_FONT"]
                alt = False
                continue

            if row.is_annotation:
                ws.cell(row=data_row_num, column=1,
                        value="  · " + row.annotation_display).font = ANNOT_FONT
                for col_offset, p in enumerate(report.periods):
                    v = row.values.get(p)
                    col_num = 4 + col_offset
                    cell = ws.cell(row=data_row_num, column=col_num)
                    if v is not None:
                        cell.value = round(v / 100, 4) if row.annotation_fmt == "percent" else round(v, 4)
                        cell.number_format = '0.0%' if row.annotation_fmt == "percent" else '0.00'
                    cell.font = ANNOT_FONT
                    cell.alignment = Alignment(horizontal="right")
                continue

            m = row.metric
            indent_str = "    " * m.indent
            label_cell = ws.cell(row=data_row_num, column=1, value=indent_str + m.display)
            ws.cell(row=data_row_num, column=2, value=m.name)
            ws.cell(row=data_row_num, column=3, value=m.unit)
            row_font = styles["TOTAL_FONT"] if m.indent == 0 else styles["DETAIL_FONT"]
            label_cell.font = row_font

            if alt:
                for c in range(1, 4 + len(report.periods)):
                    ws.cell(row=data_row_num, column=c).fill = styles["ALT_FILL"]

            for col_offset, p in enumerate(report.periods):
                v = row.values.get(p)
                col_num = 4 + col_offset
                cell = ws.cell(row=data_row_num, column=col_num)
                if v is not None:
                    if m.unit == "USD":
                        cell.value = round(v / report.scale, 2)
                        cell.number_format = '#,##0.0'
                    elif m.unit == "shares":
                        cell.value = round(v / 1_000_000, 2)
                        cell.number_format = '#,##0.0'
                    elif m.unit == "USD/shares":
                        cell.value = round(v, 2)
                        cell.number_format = '0.00'
                    else:
                        cell.value = v
                        cell.number_format = '#,##0.00'
                cell.font = row_font
                cell.alignment = Alignment(horizontal="right")
            alt = not alt

        ws.column_dimensions["A"].width = 38
        ws.column_dimensions["B"].width = 30
        ws.column_dimensions["C"].width = 12
        for col_offset in range(len(report.periods)):
            ws.column_dimensions[get_column_letter(4 + col_offset)].width = 14

    wb.save(output_path)
